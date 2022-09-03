from yahoo_fin import stock_info as si
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

print("Scraping Stocks Live Market Price...")
currentRow = 2
wb = load_workbook('stock.xlsx')
ws = wb.active
sym = ws['A']
ticker = []

print("Reading Company Code from Excel File...")
for cell in sym:
    ticker.append(cell.value)
ticker.pop(0)
print("Getting Live Stock Price from Yahoo Finance...")
for item in ticker:
    price = (si.get_live_price(item))
    ws.cell(row=currentRow, column=4).value = price
    currentRow += 1
    print(".", end='')
wb.save('stock.xlsx')
wb.close()
